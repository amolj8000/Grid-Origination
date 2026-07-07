#!/usr/bin/env python3
"""
Phase 6: GIS Report POI bus# → CIM PSSE_BUS_NUMBER → resource node → Nominatim/county geocode
Exact bus-number bridge: 104 zone_centroid nodes matched.
Run: cd artifacts/pypsa-engine && .venv/bin/python3 ../../scripts/src/geo-locate-ercot-nodes-phase6.py
"""

import os, sys, re, json, time, zipfile, csv, io, urllib.request, urllib.parse, psycopg2

DATABASE_URL = os.environ.get("DATABASE_URL", "")
CIM_ZIP  = "attached_assets/RPT.00010008.0000000000000000.20260701.001311339.CIM_Jul_ML1_1_1783382391984.zip"
GIS_XLSX = "attached_assets/RPT.00015933.0000000000000000.20260701.151514224.GIS_Report_J_1783382598151.xlsx"

# Texas county centroids (lat, lon) — 254 counties
TX_COUNTY_CENTROIDS = {
    "Anderson": (31.81, -95.65), "Andrews": (32.31, -102.64), "Angelina": (31.24, -94.61),
    "Aransas": (28.07, -97.04), "Archer": (33.62, -98.69), "Armstrong": (34.97, -101.36),
    "Atascosa": (28.89, -98.53), "Austin": (29.89, -96.27), "Bailey": (34.07, -102.83),
    "Bandera": (29.75, -99.25), "Bastrop": (30.10, -97.31), "Baylor": (33.62, -99.22),
    "Bee": (28.42, -97.74), "Bell": (31.04, -97.48), "Bexar": (29.45, -98.52),
    "Blanco": (30.27, -98.40), "Borden": (32.74, -101.43), "Bosque": (31.90, -97.64),
    "Bowie": (33.45, -94.17), "Brazoria": (29.17, -95.44), "Brazos": (30.66, -96.35),
    "Brewster": (29.83, -103.25), "Briscoe": (34.53, -100.29), "Brooks": (27.03, -98.22),
    "Brown": (31.78, -99.01), "Burleson": (30.50, -96.63), "Burnet": (30.79, -98.23),
    "Caldwell": (29.84, -97.61), "Calhoun": (28.43, -96.62), "Callahan": (32.30, -99.37),
    "Cameron": (26.14, -97.47), "Camp": (32.98, -94.97), "Carson": (35.40, -101.36),
    "Cass": (33.08, -94.35), "Castro": (34.53, -102.26), "Chambers": (29.70, -94.60),
    "Cherokee": (31.85, -95.17), "Childress": (34.53, -100.21), "Clay": (33.79, -98.21),
    "Cochran": (33.60, -102.83), "Coke": (31.89, -100.53), "Coleman": (31.83, -99.45),
    "Collin": (33.19, -96.57), "Collingsworth": (34.97, -100.27), "Colorado": (29.62, -96.52),
    "Comal": (29.83, -98.24), "Comanche": (31.95, -98.56), "Concho": (31.32, -99.87),
    "Cooke": (33.64, -97.22), "Coryell": (31.39, -97.79), "Cottle": (34.08, -100.28),
    "Crane": (31.43, -102.35), "Crockett": (30.72, -101.41), "Crosby": (33.61, -101.30),
    "Culberson": (31.45, -104.52), "Dallam": (36.28, -102.60), "Dallas": (32.77, -96.79),
    "Dawson": (32.74, -101.95), "Deaf Smith": (34.97, -102.60), "Delta": (33.39, -95.67),
    "Denton": (33.21, -97.12), "DeWitt": (29.07, -97.35), "Dickens": (33.62, -100.79),
    "Dimmit": (28.44, -99.76), "Donley": (34.97, -100.82), "Duval": (27.68, -98.52),
    "Eastland": (32.30, -98.81), "Ector": (31.87, -102.54), "Edwards": (29.98, -100.30),
    "Ellis": (32.35, -96.80), "El Paso": (31.77, -106.24), "Erath": (32.24, -98.21),
    "Falls": (31.27, -96.93), "Fannin": (33.59, -96.10), "Fayette": (29.88, -96.93),
    "Fisher": (32.74, -100.40), "Floyd": (33.97, -101.30), "Foard": (33.97, -99.78),
    "Fort Bend": (29.53, -95.77), "Franklin": (33.18, -95.22), "Freestone": (31.70, -96.15),
    "Frio": (28.87, -99.10), "Gaines": (32.74, -102.64), "Galveston": (29.24, -94.85),
    "Garza": (33.18, -101.30), "Gillespie": (30.32, -98.94), "Glasscock": (31.87, -101.52),
    "Goliad": (28.66, -97.39), "Gonzales": (29.45, -97.49), "Gray": (35.40, -100.82),
    "Grayson": (33.63, -96.68), "Gregg": (32.49, -94.82), "Grimes": (30.55, -95.98),
    "Guadalupe": (29.57, -97.94), "Hale": (34.07, -101.82), "Hall": (34.53, -100.68),
    "Hamilton": (31.70, -98.12), "Hansford": (36.28, -101.36), "Hardeman": (34.29, -99.75),
    "Hardin": (30.27, -94.38), "Harris": (29.85, -95.40), "Harrison": (32.54, -94.37),
    "Hartley": (35.84, -102.60), "Haskell": (33.17, -99.73), "Hays": (30.06, -98.03),
    "Hemphill": (35.84, -100.27), "Henderson": (32.22, -95.86), "Hidalgo": (26.39, -98.18),
    "Hill": (31.99, -97.13), "Hockley": (33.61, -102.34), "Hood": (32.44, -97.81),
    "Hopkins": (33.15, -95.57), "Houston": (31.32, -95.42), "Howard": (32.30, -101.43),
    "Hudspeth": (31.45, -105.37), "Hunt": (33.12, -95.99), "Hutchinson": (35.84, -101.36),
    "Irion": (31.30, -100.98), "Jack": (33.25, -98.17), "Jackson": (28.96, -96.57),
    "Jasper": (31.00, -93.99), "Jeff Davis": (30.72, -104.13), "Jefferson": (29.88, -94.16),
    "Jim Hogg": (27.06, -98.70), "Jim Wells": (27.73, -97.91), "Johnson": (32.38, -97.37),
    "Jones": (32.74, -99.88), "Karnes": (28.89, -97.85), "Kaufman": (32.60, -96.29),
    "Kendall": (29.94, -98.71), "Kenedy": (26.93, -97.68), "Kent": (33.18, -100.78),
    "Kerr": (30.06, -99.35), "Kimble": (30.49, -99.75), "King": (33.62, -100.26),
    "Kinney": (29.35, -100.42), "Kleberg": (27.43, -97.70), "Knox": (33.61, -99.74),
    "Lamar": (33.67, -95.58), "Lamb": (34.07, -102.34), "Lampasas": (31.19, -98.24),
    "La Salle": (28.35, -99.10), "Lavaca": (29.38, -96.93), "Lee": (30.32, -96.97),
    "Leon": (31.30, -95.98), "Liberty": (30.15, -94.83), "Limestone": (31.54, -96.59),
    "Lipscomb": (36.28, -100.27), "Live Oak": (28.35, -98.10), "Llano": (30.72, -98.69),
    "Loving": (31.85, -103.63), "Lubbock": (33.61, -101.82), "Lynn": (33.18, -101.82),
    "McCulloch": (31.20, -99.35), "McLennan": (31.55, -97.20), "McMullen": (28.36, -98.57),
    "Madison": (30.96, -95.92), "Marion": (32.82, -94.37), "Martin": (32.30, -101.95),
    "Mason": (30.72, -99.23), "Matagorda": (28.79, -96.02), "Maverick": (28.75, -100.31),
    "Medina": (29.35, -99.11), "Menard": (30.88, -99.82), "Midland": (31.87, -102.03),
    "Milam": (30.79, -96.98), "Mills": (31.50, -98.60), "Mitchell": (32.30, -100.92),
    "Montague": (33.67, -97.72), "Montgomery": (30.30, -95.51), "Moore": (35.84, -101.90),
    "Morris": (33.10, -94.73), "Motley": (34.07, -100.78), "Nacogdoches": (31.61, -94.62),
    "Navarro": (32.05, -96.48), "Newton": (30.78, -93.75), "Nolan": (32.30, -100.40),
    "Nueces": (27.73, -97.56), "Ochiltree": (36.28, -100.82), "Oldham": (35.40, -102.60),
    "Orange": (30.12, -93.90), "Palo Pinto": (32.75, -98.31), "Panola": (32.16, -94.31),
    "Parker": (32.78, -97.80), "Parmer": (34.53, -102.84), "Pecos": (30.79, -102.72),
    "Polk": (30.79, -94.84), "Potter": (35.40, -101.90), "Presidio": (29.87, -104.28),
    "Rains": (32.87, -95.79), "Randall": (34.97, -101.90), "Reagan": (31.37, -101.52),
    "Real": (29.83, -99.83), "Red River": (33.62, -94.97), "Reeves": (31.32, -103.69),
    "Refugio": (28.43, -97.16), "Roberts": (35.84, -100.82), "Robertson": (31.03, -96.51),
    "Rockwall": (32.90, -96.42), "Runnels": (31.83, -99.97), "Rusk": (32.10, -94.77),
    "Sabine": (31.34, -93.86), "San Augustine": (31.39, -94.17), "San Jacinto": (30.58, -95.17),
    "San Patricio": (27.97, -97.52), "San Saba": (31.15, -98.72), "Schleicher": (30.90, -100.54),
    "Scurry": (32.74, -100.92), "Shackelford": (32.74, -99.35), "Shelby": (31.79, -94.14),
    "Sherman": (36.28, -101.90), "Smith": (32.38, -95.27), "Somervell": (32.22, -97.77),
    "Starr": (26.56, -98.74), "Stephens": (32.74, -98.81), "Sterling": (31.83, -101.05),
    "Stonewall": (33.18, -99.88), "Sutton": (30.49, -100.54), "Swisher": (34.53, -101.74),
    "Tarrant": (32.77, -97.29), "Taylor": (32.30, -99.88), "Terrell": (30.22, -102.08),
    "Terry": (33.17, -102.34), "Throckmorton": (33.18, -99.22), "Titus": (33.22, -94.97),
    "Tom Green": (31.42, -100.44), "Travis": (30.33, -97.78), "Trinity": (31.09, -95.14),
    "Tyler": (30.78, -94.38), "Upshur": (32.73, -94.97), "Upton": (31.37, -102.03),
    "Uvalde": (29.35, -99.79), "Val Verde": (29.88, -101.16), "Van Zandt": (32.56, -95.86),
    "Victoria": (28.79, -97.00), "Walker": (30.73, -95.57), "Waller": (29.97, -95.99),
    "Ward": (31.51, -103.11), "Washington": (30.21, -96.40), "Webb": (27.76, -99.45),
    "Wharton": (29.28, -96.23), "Wheeler": (35.40, -100.27), "Wichita": (33.99, -98.69),
    "Wilbarger": (34.08, -99.24), "Willacy": (26.46, -97.68), "Williamson": (30.65, -97.60),
    "Wilson": (29.17, -98.09), "Winkler": (31.83, -103.06), "Wise": (33.21, -97.66),
    "Wood": (32.78, -95.37), "Yoakum": (33.17, -102.83), "Young": (33.18, -98.69),
    "Zapata": (27.00, -99.19), "Zavala": (28.87, -99.76),
    # Alternate spellings
    "Deaf smith": (34.97, -102.60), "Jim wells": (27.73, -97.91), "Jim hogg": (27.06, -98.70),
    "La salle": (28.35, -99.10), "Live oak": (28.35, -98.10), "Mc culloch": (31.20, -99.35),
    "Mc mullen": (28.36, -98.57), "Tom green": (31.42, -100.44), "Val verde": (29.88, -101.16),
    "Van zandt": (32.56, -95.86),
}

def nominatim_geocode(query, county, delay=1.1):
    """Try Nominatim geocoding; validate result is within Texas."""
    try:
        q = urllib.parse.quote_plus(query + " Texas")
        url = f"https://nominatim.openstreetmap.org/search?q={q}&format=json&limit=3&countrycodes=us"
        req = urllib.request.Request(url, headers={"User-Agent": "grid-platform/1.0"})
        with urllib.request.urlopen(req, timeout=8) as resp:
            results = json.loads(resp.read())
        time.sleep(delay)  # Nominatim rate limit: 1 req/sec
        for r in results:
            lat, lon = float(r["lat"]), float(r["lon"])
            # Validate: must be within Texas bounds
            if 25.8 <= lat <= 36.5 and -106.7 <= lon <= -93.5:
                return lat, lon
    except Exception as e:
        pass
    return None, None

def county_centroid(county_name):
    cn = county_name.strip().title()
    if cn in TX_COUNTY_CENTROIDS:
        return TX_COUNTY_CENTROIDS[cn]
    cn2 = cn.replace(" County","").strip()
    return TX_COUNTY_CENTROIDS.get(cn2, (None, None))

def best_project(projs):
    """Pick the most representative project from a set sharing a bus."""
    # Prefer projects with clean sub_name (no 'tap', 'LOCAL', numbers)
    def score(p):
        sn = p["sub_name"]
        s = 0
        if re.match(r'^[A-Za-z]', sn): s += 3
        if len(sn) > 4: s += 2
        if p["county"]: s += 2
        if re.search(r'\d{4,}', sn): s -= 2  # raw bus numbers in name
        if "LOCAL" in sn.upper() or "TAP" in sn.upper(): s -= 2
        return s
    return max(projs, key=score)

def main():
    conn = psycopg2.connect(DATABASE_URL)
    cur = conn.cursor()

    # Load zone_centroid nodes
    cur.execute("""
        SELECT node_name, substation, load_zone FROM ercot_node_locations
        WHERE node_type='resource_node' AND location_source='zone_centroid'
        ORDER BY node_name
    """)
    zc_nodes = {r[0]: {"substation": r[1], "zone": r[2]} for r in cur.fetchall()}
    print(f"Zone-centroid nodes: {len(zc_nodes)}")

    # Load CIM PSSE_BUS_NUMBER map
    print("Loading CIM Settlement_Points...")
    cim_zip = zipfile.ZipFile(CIM_ZIP)
    with cim_zip.open("SP_List_EB_Mapping/Settlement_Points_06182026_102348.csv") as f:
        sp_rows = [r for r in csv.DictReader(io.StringIO(f.read().decode("utf-8", errors="replace")))
                   if r.get("RESOURCE_NODE", "").strip()]
    rn_to_bus = {r["RESOURCE_NODE"].strip(): r["PSSE_BUS_NUMBER"].strip() for r in sp_rows if r["PSSE_BUS_NUMBER"].strip()}
    print(f"CIM resource nodes with PSSE bus: {len(rn_to_bus)}")

    # Parse GIS Report
    print("Loading GIS Report...")
    import openpyxl
    wb = openpyxl.load_workbook(GIS_XLSX, read_only=False, data_only=True)
    ws = wb["Project Details - Large Gen"]
    hdr_row = ws[31]
    headers = {c.column: c.value for c in hdr_row if c.value}
    bus_to_proj = {}
    for row_idx in range(36, ws.max_row + 1):
        row = ws[row_idx]
        vals = {headers.get(c.column, "?"): c.value for c in row if c.value is not None and c.column in headers}
        if not vals.get("INR"):
            continue
        poi = str(vals.get("POI Location", "") or "")
        bus_num = ""
        sub_name = poi
        m = re.match(r"^(\d{3,6})\s+(.+?)(?:\s+\d+\s*kV\s*)?$", poi.strip())
        if m:
            bus_num = m.group(1)
            sub_name = re.sub(r"\s*\d+\s*kV\s*$", "", m.group(2)).strip()
        tap = re.search(r"tap\s+\d+kV\s+(\d{3,6})\s+(.+)", poi, re.IGNORECASE)
        if tap:
            bus_num = tap.group(1)
            sub_name = re.sub(r"\s*\d+\s*kV\s*$", "", tap.group(2)).strip()
            sub_name = sub_name.split("–")[0].split("-")[0].strip()
        if not bus_num:
            continue
        proj = {
            "name": str(vals.get("Project Name", "")).strip(),
            "bus_num": bus_num,
            "sub_name": sub_name,
            "county": str(vals.get("County", "")).strip(),
            "zone": str(vals.get("CDR Reporting Zone", "")).strip(),
            "fuel": str(vals.get("Fuel", "")).strip(),
        }
        if bus_num not in bus_to_proj:
            bus_to_proj[bus_num] = []
        bus_to_proj[bus_num].append(proj)
    wb.close()
    print(f"GIS projects with bus#: {sum(len(v) for v in bus_to_proj.values())} across {len(bus_to_proj)} unique buses")

    # Phase 6: match zone_centroid nodes via bus number
    updates = []
    no_cim_bus = 0
    no_gis_match = 0

    for node_name, info in sorted(zc_nodes.items()):
        bus = rn_to_bus.get(node_name, "")
        if not bus:
            no_cim_bus += 1
            continue
        projs = bus_to_proj.get(bus, [])
        if not projs:
            no_gis_match += 1
            continue
        proj = best_project(projs)
        updates.append({
            "node": node_name,
            "bus": bus,
            "proj": proj,
            "zone": info["zone"],
        })

    unique_nodes = len(set(u["node"] for u in updates))
    print(f"\nBus-matched nodes: {unique_nodes} unique (from {len(updates)} project matches)")
    print(f"No CIM bus: {no_cim_bus}  No GIS match: {no_gis_match}")

    # Deduplicate to one best project per node
    best_per_node = {}
    for u in updates:
        nd = u["node"]
        if nd not in best_per_node:
            best_per_node[nd] = u
        # (already picked best_project per bus, so just take first)

    print(f"\nNodes to geocode: {len(best_per_node)}")
    print("\nGeocoding via Nominatim (1 req/sec) then county centroid fallback...")
    print("-" * 70)

    geocoded = []
    nominatim_hits = 0
    county_hits = 0
    failed = 0

    for i, (node_name, u) in enumerate(sorted(best_per_node.items())):
        proj = u["proj"]
        sub_name = proj["sub_name"]
        county = proj["county"]
        zone = u["zone"]

        # Try Nominatim with substation name
        lat, lon = None, None
        source = None
        eia_name = None

        # Only Nominatim if sub_name looks like a real place (not an ERCOT code)
        sub_clean = re.sub(r"[^a-zA-Z\s]", "", sub_name).strip()
        looks_like_place = (
            len(sub_clean) >= 5 and
            not re.match(r"^[A-Z]{2,5}\d", sub_name) and  # skip pure ERCOT codes like "BYRSW_5"
            not sub_name.startswith("Bus") and
            not "LOCAL" in sub_name.upper()
        )
        if looks_like_place:
            query = f"{sub_clean} {county}"
            lat, lon = nominatim_geocode(query, county)
            if lat:
                source = "nominatim_poi"
                eia_name = f"{sub_name} ({county} Co) [POI bus {proj['bus_num']}]"
                nominatim_hits += 1
                print(f"  [{i+1:3d}] NOMINATIM {node_name:<25} sub={sub_clean[:20]:<20} → ({lat:.4f},{lon:.4f})")

        if not lat:
            # County centroid fallback
            clat, clon = county_centroid(county)
            if clat:
                lat, lon = clat, clon
                source = "county_centroid"
                eia_name = f"{proj['name'][:50]} (bus {proj['bus_num']}, {county} Co)"
                county_hits += 1
                if i < 20 or i % 20 == 0:
                    print(f"  [{i+1:3d}] county   {node_name:<25} county={county:<15} → ({lat:.4f},{lon:.4f})")
            else:
                failed += 1
                print(f"  [{i+1:3d}] FAILED   {node_name:<25} county={county} (no centroid)")

        if lat:
            geocoded.append({
                "node": node_name,
                "lat": lat,
                "lon": lon,
                "source": source,
                "eia_name": eia_name,
                "bus": proj["bus_num"],
            })

    print(f"\nResults: Nominatim={nominatim_hits} county={county_hits} failed={failed}")
    print(f"Geocoded: {len(geocoded)} nodes ready to apply")

    if not geocoded:
        print("No updates to apply.")
        cur.close(); conn.close()
        return

    print(f"\nApply {len(geocoded)} updates? [y/N]: ", end="")
    ans = input().strip().lower()
    if ans != "y":
        print("Aborted.")
        cur.close(); conn.close()
        return

    applied = 0
    for u in geocoded:
        cur.execute("""
            UPDATE ercot_node_locations
            SET latitude=%s, longitude=%s, location_source=%s, eia_plant_name=%s
            WHERE node_name=%s AND node_type='resource_node'
              AND location_source='zone_centroid'
        """, (u["lat"], u["lon"], u["source"], u["eia_name"], u["node"]))
        if cur.rowcount > 0:
            applied += 1
    conn.commit()
    print(f"Applied {applied} updates.")

    cur.execute("""
        SELECT location_source, COUNT(*) FROM ercot_node_locations
        WHERE node_type='resource_node' GROUP BY location_source ORDER BY count DESC
    """)
    print("\nFinal breakdown:")
    for r in cur.fetchall():
        print(f"  {r[0]}: {r[1]}")

    cur.close()
    conn.close()

if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""
Phase 5: CIM unit_sub → EIA-860 + Queue fuzzy match
Uses UNIT_SUBSTATION from CDR 10008 Resource_Node_to_Unit CSV
(better than node prefix for short-name nodes like ANG_ALL, ALP_BESS_RN, AMO_AMOCO_*)
Run: cd artifacts/pypsa-engine && .venv/bin/python3 ../../scripts/src/geo-locate-ercot-nodes-phase5.py
"""

import os, sys, re, zipfile, csv, io, json, psycopg2
from rapidfuzz import fuzz

DATABASE_URL = os.environ.get("DATABASE_URL", "")
CIM_ZIP = "attached_assets/RPT.00010008.0000000000000000.20260701.001311339.CIM_Jul_ML1_1_1783382391984.zip"
EIA_ZIP = "attached_assets/eia8602024_1777780153233.zip"

STOP_WORDS = {
    "wind","solar","farm","project","llc","inc","corp","energy","power","electric",
    "generation","station","plant","facility","battery","storage","bess","ess",
    "resources","resource","holdings","partners","ventures","texas","tx",
    "all","rn","gen","unit","units","the","and","of","at",
}

TECH_TOKENS = {"slr","wnd","wind","ess","bess","all","rn","gen","gt","ct","st"}

def clean_name(s):
    s = re.sub(r"[^a-z0-9\s]", " ", s.lower())
    tokens = [t for t in s.split() if t not in STOP_WORDS and len(t) > 1]
    return " ".join(tokens).strip()

def clean_unit_sub(s):
    """Clean CIM unit substation code into matchable tokens."""
    s = s.lower()
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    tokens = [t for t in s.split() if t not in TECH_TOKENS and len(t) > 1]
    return " ".join(tokens).strip()

def score(a, b):
    if not a or not b:
        return 0
    nc = len(a.replace(" ", ""))
    if nc < 4:
        return 0
    if nc <= 6:
        return fuzz.token_sort_ratio(a, b) * 0.5 + fuzz.ratio(a, b) * 0.5
    return fuzz.partial_ratio(a, b) * 0.5 + fuzz.token_sort_ratio(a, b) * 0.3 + fuzz.ratio(a, b) * 0.2

def main():
    conn = psycopg2.connect(DATABASE_URL)
    cur = conn.cursor()

    # 1. Load zone_centroid nodes from DB
    cur.execute("""
        SELECT node_name, substation, load_zone
        FROM ercot_node_locations
        WHERE node_type='resource_node' AND location_source='zone_centroid'
        ORDER BY node_name
    """)
    zc_nodes = {r[0]: {"substation": r[1], "zone": r[2]} for r in cur.fetchall()}
    print(f"Zone-centroid nodes to improve: {len(zc_nodes)}")

    # 2. Load CIM unit_sub map
    print("Loading CIM Resource_Node_to_Unit...")
    cim_zip = zipfile.ZipFile(CIM_ZIP)
    with cim_zip.open("SP_List_EB_Mapping/Resource_Node_to_Unit_06182026_102348.csv") as f:
        rnu_rows = list(csv.DictReader(io.StringIO(f.read().decode("utf-8", errors="replace"))))
    cim_usub = {}
    for r in rnu_rows:
        rn = r["RESOURCE_NODE"].strip()
        us = r["UNIT_SUBSTATION"].strip()
        if rn not in cim_usub:
            cim_usub[rn] = us
    print(f"CIM unit_sub entries: {len(cim_usub)}")

    # 3. Load EIA-860 TX plant data
    print("Loading EIA-860 plants...")
    eia_zip = zipfile.ZipFile(EIA_ZIP)
    plant_file = next(n for n in eia_zip.namelist() if "Plant_Y2024" in n and n.endswith(".xlsx"))
    import openpyxl
    with eia_zip.open(plant_file) as f:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value) if c.value else "" for c in next(ws.iter_rows(min_row=2, max_row=2))]
        name_col = headers.index("Plant Name") if "Plant Name" in headers else None
        state_col = headers.index("State") if "State" in headers else None
        lat_col = headers.index("Latitude") if "Latitude" in headers else None
        lon_col = headers.index("Longitude") if "Longitude" in headers else None
        print(f"EIA cols: name={name_col}, state={state_col}, lat={lat_col}, lon={lon_col}")

        eia_plants = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            if state_col is not None and row[state_col] != "TX":
                continue
            name = row[name_col] if name_col is not None else None
            lat = row[lat_col] if lat_col is not None else None
            lon = row[lon_col] if lon_col is not None else None
            if name and lat and lon:
                try:
                    eia_plants.append({"name": str(name), "clean": clean_name(str(name)), "lat": float(lat), "lon": float(lon)})
                except (ValueError, TypeError):
                    pass
        wb.close()
    print(f"EIA TX plants with lat/lon: {len(eia_plants)}")

    # 4. Load queue projects (ERCOT, with lat/lon)
    cur.execute("""
        SELECT project_name, latitude, longitude
        FROM queue_projects
        WHERE market='ERCOT' AND latitude IS NOT NULL AND longitude IS NOT NULL
    """)
    queue_pts = [{"name": r[0], "clean": clean_name(str(r[0])), "lat": float(r[1]), "lon": float(r[2])}
                 for r in cur.fetchall() if r[0]]
    print(f"Queue projects with lat/lon: {len(queue_pts)}")

    # 5. For each zone_centroid node, get unit_sub and try fuzzy match
    updates = []
    skipped_no_cim = 0
    skipped_short = 0
    tried = 0

    for node_name, info in sorted(zc_nodes.items()):
        unit_sub = cim_usub.get(node_name, "")
        if not unit_sub:
            skipped_no_cim += 1
            continue

        # Build search key from unit_sub (primary) + node prefix (secondary)
        usub_clean = clean_unit_sub(unit_sub)
        prefix = node_name.split("_")[0]
        prefix_clean = clean_unit_sub(prefix)

        # Use unit_sub if it's longer/better; else prefix
        if len(usub_clean) >= len(prefix_clean):
            search_key = usub_clean
        else:
            search_key = prefix_clean

        if len(search_key.replace(" ", "")) < 4:
            skipped_short += 1
            continue

        tried += 1

        # Phase 5a: EIA-860 plant match using unit_sub
        best_eia = None
        best_eia_score = 0
        for plant in eia_plants:
            s = score(search_key, plant["clean"])
            if s > best_eia_score:
                best_eia_score = s
                best_eia = plant

        # Phase 5b: Queue match using unit_sub
        best_q = None
        best_q_score = 0
        for pt in queue_pts:
            s = score(search_key, pt["clean"])
            if s > best_q_score:
                best_q_score = s
                best_q = pt

        EIA_THRESH = 82
        Q_THRESH = 77

        def token_present(key, plant_name):
            """At least one non-trivial token from key appears in plant_name."""
            key_tokens = [t for t in key.split() if len(t) >= 3]
            pn_lower = plant_name.lower()
            return any(t in pn_lower for t in key_tokens)

        if best_eia_score >= EIA_THRESH and token_present(search_key, best_eia["name"]):
            updates.append({
                "node": node_name,
                "lat": best_eia["lat"],
                "lon": best_eia["lon"],
                "source": "eia_fuzzy_match",
                "eia_name": best_eia["name"],
                "score": best_eia_score,
                "key": search_key,
                "orig_name": unit_sub,
            })
        elif best_q_score >= Q_THRESH and token_present(search_key, best_q["name"]):
            updates.append({
                "node": node_name,
                "lat": best_q["lat"],
                "lon": best_q["lon"],
                "source": "queue_latlon_match",
                "eia_name": best_q["name"][:80],
                "score": best_q_score,
                "key": search_key,
                "orig_name": unit_sub,
            })

    print(f"\nResults:")
    print(f"  Tried: {tried}")
    print(f"  Skipped (no CIM entry): {skipped_no_cim}")
    print(f"  Skipped (too short): {skipped_short}")
    print(f"  Matches found: {len(updates)}")
    print()
    print("Top matches:")
    for u in sorted(updates, key=lambda x: -x["score"])[:20]:
        print(f"  {u['node']:<25} unit_sub={u['orig_name']:<15} key={u['key']:<18} → {u['eia_name'][:35]:<35} ({u['source'][:8]}, {u['score']:.1f})")

    if not updates:
        print("No new matches found.")
        cur.close(); conn.close()
        return

    print(f"\nApply {len(updates)} updates? [y/N]: ", end="")
    ans = input().strip().lower()
    if ans != "y":
        print("Aborted.")
        cur.close(); conn.close()
        return

    applied = 0
    for u in updates:
        cur.execute("""
            UPDATE ercot_node_locations
            SET latitude=%s, longitude=%s, location_source=%s, eia_plant_name=%s
            WHERE node_name=%s AND node_type='resource_node'
              AND location_source='zone_centroid'
        """, (u["lat"], u["lon"], u["source"], u["eia_name"], u["node"]))
        if cur.rowcount > 0:
            applied += 1

    conn.commit()
    print(f"Applied {applied} updates to DB.")

    # Summary
    cur.execute("""
        SELECT location_source, COUNT(*) FROM ercot_node_locations
        WHERE node_type='resource_node' GROUP BY location_source ORDER BY count DESC
    """)
    print("\nFinal breakdown:")
    for r in cur.fetchall():
        print(f"  {r[0]}: {r[1]}")

    cur.close()
    conn.close()

if __name__ == "__main__":
    main()
